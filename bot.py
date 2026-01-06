from telegram import Update
from telegram.ext import Application, MessageHandler, CommandHandler, filters, ContextTypes
import openpyxl
import re
import os

TOKEN = os.environ.get('BOT_TOKEN')
EXCEL_FILE = "price.xlsx"

print("🤖 Запуск бота...")

# Загрузка прайс-листа
wb = openpyxl.load_workbook(EXCEL_FILE)
sheet = wb.active

# Найти колонки РВ-1 и РВр-1
rv1_col = rvr1_col = None
for col in range(1, sheet.max_column + 1):
    val = str(sheet.cell(1, col).value or '').lower().strip()
    if val == 'рв-1' and not rv1_col:
        rv1_col = col
    if val == 'рвр-1' and not rvr1_col:
        rvr1_col = col

print(f"✅ РВ-1: колонка {rv1_col}, РВр-1: колонка {rvr1_col}")

def find_price(grill_type, height, width):
    start_col = rv1_col if grill_type == 'РВ-1' else rvr1_col
    
    for row in range(3, sheet.max_row + 1):
        h = sheet.cell(row, start_col).value
        w = sheet.cell(row, start_col + 1).value
        
        if h == height and w == width:
            price = sheet.cell(row, start_col + 2).value
            return float(price) if price else None
    return None

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет! Я бот для поиска цен на решетки.\n\n"
        "📝 Формат:\n"
        "• РВ-1 100(h)х200\n"
        "• РВр-1 300(h)х300\n\n"
        "Просто напишите размер!"
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    match = re.match(r'(РВ-1|РВр-1)\s+(\d+)(?:\(h\)|h)?\s*[хx]\s*(\d+)', text, re.I)
    
    if not match:
        await update.message.reply_text("❌ Неверный формат!\n\nИспользуйте: РВ-1 100(h)х200")
        return
    
    grill_type, height, width = match.groups()
    height, width = int(height), int(width)
    price = find_price(grill_type, height, width)
    
    if not price:
        await update.message.reply_text(f"❌ Цена не найдена для {grill_type} {height}(h)х{width}")
        return
    
    response = f"""✅ {grill_type} {height}(h)х{width}

💰 Розничная: {price:,.0f} ₸

🎯 Скидка 10%: {price * 0.9:,.0f} ₸
🎯 Скидка 20%: {price * 0.8:,.0f} ₸
🎯 Скидка 30%: {price * 0.7:,.0f} ₸"""
    
    await update.message.reply_text(response)

def main():
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    print("✅ Бот запущен!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
```

4. Нажмите **"Commit changes"** внизу

---

### Файл 2: `requirements.txt`

1. Снова нажмите **"Add file"** → **"Create new file"**
2. Назовите: `requirements.txt`
3. Вставьте:
```
python-telegram-bot==20.7
openpyxl==3.1.2
```

4. Нажмите **"Commit changes"**

---

### Файл 3: Загрузите Excel файл

1. Нажмите **"Add file"** → **"Upload files"**
2. Перетащите ваш Excel файл с прайс-листом
3. **Переименуйте** его в `price.xlsx` (важно!)
4. Нажмите **"Commit changes"**

---

## 🚀 Шаг 5: Подключите Render

1. Перейдите на [render.com](https://render.com)
2. Нажмите **"Get Started"** → зарегистрируйтесь через **GitHub**
3. После входа нажмите **"New +"** → **"Background Worker"**
4. Нажмите **"Connect account"** чтобы подключить GitHub
5. Выберите репозиторий **telegram-price-bot**
6. Настройки:
   - **Name**: telegram-price-bot
   - **Region**: Frankfurt (или ближайший к вам)
   - **Branch**: main
   - **Build Command**: `pip install -r requirements.txt`
   - **Start Command**: `python bot.py`
   - **Plan**: Free

---

## 🔐 Шаг 6: Добавьте токен

1. Прокрутите вниз до раздела **"Environment Variables"**
2. Нажмите **"Add Environment Variable"**
3. **Key**: `BOT_TOKEN`
4. **Value**: вставьте ваш токен от BotFather
5. Нажмите **"Add"**

---

## ✅ Шаг 7: Запустите!

1. Нажмите **"Create Background Worker"** внизу
2. Подождите 2-3 минуты (Render устанавливает всё)
3. Вы увидите логи:
```
✅ РВ-1: колонка X, РВр-1: колонка Y
✅ Бот запущен!
